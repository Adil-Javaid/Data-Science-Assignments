# 18th September 2023
# CSC461 – Assignment1 – Web Scraping 
# H. M. Adil Javaid
# FA21-BSE-006
# In this task get the movies title name's and rating and then save data into excel file

import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import emoji


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
}

url = ["https://www.imdb.com/title/tt0468569/?ref_=chttp_t_3","https://www.imdb.com/title/tt0113277/?ref_=chttp_t_111","https://www.imdb.com/title/tt5074352/?ref_=chttp_t_126",
       "https://www.imdb.com/title/tt0469494/?ref_=chttp_t_138","https://www.imdb.com/title/tt0012349/?ref_=chttp_t_131"]

excel = openpyxl.Workbook()
worksheet = excel.active

worksheet['A1'] = 'Title'
worksheet['B1'] = 'Rating' + emoji.emojize(":star:")

row = 2
for i in url:
    req = requests.get(i,headers=headers)
    soup = BeautifulSoup(req.content,"html.parser")
        
    title_element = soup.find("h1", {"class": "sc-afe43def-0 hnYaOZ"})
    title = title_element.text.strip()

    rating_element = soup.find("span", {"class": "sc-bde20123-1 iZlgcd"})
    rating = float(rating_element.text.strip())

    worksheet.cell(row = row, column =1, value = title)
    worksheet.cell(row = row, column =2, value = rating)
    row += 1
    time.sleep(1)

excel.save('Movies_Title_Rating.xlsx')